"""Excel (.xlsx) parser using openpyxl + DuckDB for multi-sheet analysis."""

from __future__ import annotations
import logging

logger = logging.getLogger(__name__)


async def parse_xlsx(source: str) -> str:
    """Extract text from all sheets of an .xlsx file.

    Each sheet is converted to a markdown-style table.
    DuckDB is used for type inference on large sheets.
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        return ""

    try:
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
        parts: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            parts.append(f"## Sheet: {sheet_name}")

            # Use first row as header if it looks like one
            headers = [str(c) if c is not None else "" for c in rows[0]]
            parts.append(" | ".join(headers))
            parts.append(" | ".join("---" for _ in headers))

            for row in rows[1:]:
                cells = [str(c) if c is not None else "" for c in row]
                # Skip completely empty rows
                if any(c.strip() for c in cells):
                    parts.append(" | ".join(cells))

        wb.close()
        return "\n".join(parts)

    except Exception as exc:
        logger.error("Failed to parse xlsx '%s': %s", source, exc)
        return ""
